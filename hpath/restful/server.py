"""The simulation server.

REST API
--------

..
  #pylint: disable=line-too-long

+---------------------------------------+-----------------+------------------------------------------------------+
| **Endpoint**                          | **HTTP Method** | **Python function signature**                        |
+=======================================+=================+======================================================+
| ``/``                                 | GET             | :py:func:`~hpath.restful.server.hello_world()`       |
+                                       +-----------------+------------------------------------------------------+
|                                       | DELETE          | :py:func:`~hpath.restful.server.reset()`             |
+---------------------------------------+-----------------+------------------------------------------------------+
| ``/scenarios/``                       | POST            | :py:func:`~hpath.restful.server.new_scenario_rest`   |
|                                       +-----------------+------------------------------------------------------+
|                                       | GET             | :py:func:`~hpath.restful.server.list_scenarios_rest` |
+---------------------------------------+-----------------+------------------------------------------------------+
| ``/scenarios/<scenario_id>/status/``  | GET             | :py:func:`~hpath.restful.server.status_rest`         |
+---------------------------------------+-----------------+------------------------------------------------------+
| ``/scenarios/<scenario_id>/results/`` | GET             | :py:func:`~hpath.restful.server.results_rest`        |
+---------------------------------------+-----------------+------------------------------------------------------+
| ``/multi/``                           | POST            | :py:func:`~hpath.restful.server.new_multi_rest`      |
|                                       +-----------------+------------------------------------------------------+
|                                       | GET             | :py:func:`~hpath.restful.server.list_multis_rest`    |
+---------------------------------------+-----------------+------------------------------------------------------+
| ``/multi/<analysis_id>/status/``      | GET             | :py:func:`~hpath.restful.server.status_multi_rest`   |
+---------------------------------------+-----------------+------------------------------------------------------+
| ``/multi/<analysis_id>/results/``     | GET             | :py:func:`~hpath.restful.server.results_multi_rest`  |
+---------------------------------------+-----------------+------------------------------------------------------+

..
  #pylint: enable=line-too-long
"""

import json
import sqlite3 as sql
from http import HTTPStatus
from math import isnan
from typing import Any

import flask
import pandas as pd
from flask import Flask, Response, request
from flask_cors import CORS
from openpyxl import load_workbook
from werkzeug.exceptions import HTTPException

from ..conf import DB_PATH, BACKEND_PORT
from ..config import Config
from ..kpis import multi_mean_tats, multi_mean_util, multi_util_hourlies
from ..simulate import simulate
from .redis import REDIS_QUEUE

app = Flask(__name__)
cors = CORS(app, resources={r"/*": {"origins": "*"}})

SQL_CREATE_SCENARIOS = """\
CREATE TABLE IF NOT EXISTS scenarios (
    scenario_id INTEGER PRIMARY KEY AUTOINCREMENT,
    analysis_id INTEGER,
    created REAL NOT NULL,
    completed REAL,
    num_reps INTEGER NOT NULL,
    done_reps INTEGER NOT NULL,
    results TEXT,
FOREIGN KEY (analysis_id)
    REFERENCES multis (analysis_id)
)"""

SQL_CREATE_MULTIS = """\
CREATE TABLE IF NOT EXISTS multis (
    analysis_id INTEGER PRIMARY KEY AUTOINCREMENT
)
"""

SQL_INSERT_SCENARIO = """\
INSERT INTO scenarios (analysis_id, created, num_reps, done_reps)
VALUES (?,?,?,?)"""

SQL_SELECT_SCENARIO = """\
SELECT scenario_id, analysis_id, num_reps, CAST(done_reps AS REAL)/num_reps AS progress,
    created, completed
FROM scenarios
WHERE scenario_id = ?"""  # with or without analysis_id (can be None)

SQL_LIST_SCENARIOS = """\
SELECT scenario_id, analysis_id, num_reps, CAST(done_reps AS REAL)/num_reps AS progress,
    created, completed
FROM scenarios"""

SQL_SCENARIO_RESULTS = """SELECT results FROM scenarios WHERE scenario_id = ?"""

SQL_SELECT_MULTI = """\
SELECT
    analysis_id,
    GROUP_CONCAT(scenario_id) as scenario_ids,
    MIN(created) as created,
    CASE WHEN COUNT(completed) = COUNT(*) THEN MAX(completed) END as completed,
    CAST(SUM(done_reps) AS REAL) / SUM(num_reps) AS progress
FROM scenarios
WHERE analysis_id = ?
GROUP BY analysis_id
"""

SQL_LIST_MULTIS = """\
SELECT
    analysis_id,
    GROUP_CONCAT(scenario_id) as scenario_ids,
    MIN(created) as created,
    CASE WHEN COUNT(completed) = COUNT(*) THEN MAX(completed) END as completed,
    CAST(SUM(done_reps) AS REAL) / SUM(num_reps) AS progress
FROM scenarios
WHERE analysis_id IS NOT NULL
GROUP BY analysis_id
"""


@app.errorhandler(HTTPException)
def handle_exception(exc: HTTPException):
    """Return JSON instead of HTML (the default) for HTTP errors."""
    # start with the correct headers and status code from the error
    response = exc.get_response()
    # replace the body with JSON
    response.data = json.dumps({
        "code": exc.code,
        "name": exc.name,
        "description": exc.description,
    }, separators=(',', ':'))
    response.content_type = "application/json"
    return response


@app.route('/')
def hello_world() -> Response:
    """Return a simple HTML message, unless the request body is 'PING' (returns 'PONG' instead.)"""
    if request.get_data(as_text=True) == 'PING':
        return Response('PONG')
    return Response("<h1>Hello World!</h1>", status=HTTPStatus.OK)


@app.route('/', methods=['DELETE'])
def reset() -> Response:
    """Reset the database."""
    conn = sql.connect(DB_PATH)
    cur = conn.cursor()
    cur.execute("DROP TABLE multis")
    cur.execute("DROP TABLE scenarios")
    cur.execute(SQL_CREATE_MULTIS)
    cur.execute(SQL_CREATE_SCENARIOS)
    conn.commit()
    return Response(None, status=HTTPStatus.OK)


def new_scenario(config: Config) -> dict[str, Any]:
    """Set up a new simulation task from an :py:class:`Config` and submit it to the RQ server.
    This :py:class:`Config` is created from an Excel file in :py:meth:`new_scenario_rest`."""
    conn = sql.connect(DB_PATH)
    cur = conn.cursor()
    cur.execute(
        SQL_INSERT_SCENARIO,
        (config.analysis_id, config.created, config.num_reps, 0)
    )
    conn.commit()
    scenario_id = cur.lastrowid

    # Enqueue the simulation job
    REDIS_QUEUE.enqueue(simulate, config, scenario_id)

    ret = {
        'scenario_id': scenario_id,
        'created': config.created
    }
    if config.analysis_id is not None:
        ret['analysis_id'] = config.analysis_id
    return ret


@app.route('/scenarios/', methods=['POST'])
@app.route('/scenarios', methods=['POST'])
def new_scenario_rest() -> Response:
    """Process POST request for creating a new scenario.  The scenario configuration is contained
    in the request's ``files`` and ``form`` data.

    Invalid configuration files will lead to an ``HTTP 400 Bad Request`` error.
    """
    if 'config' not in request.files:
        flask.abort(HTTPStatus.BAD_REQUEST, 'Missing "config" in request.files.')
    file = request.files['config']  # Uploaded config file

    if 'num_reps' not in request.form:
        flask.abort(HTTPStatus.BAD_REQUEST, 'Missing "num_reps" in request.form.')
    num_reps = int(request.form['num_reps'])  # Number of simulation replications

    if 'sim_hours' not in request.form:
        flask.abort(HTTPStatus.BAD_REQUEST, 'Missing "sim_hours" in request.form.')
    sim_hours = float(request.form['sim_hours'])  # Simulation duration in hours

    # Parse and validate the input file. If error, return HTTP 400 Bad Request.
    try:
        config_bytes = file.stream
        wbook = load_workbook(config_bytes, data_only=True)
    except Exception as exc:
        flask.abort(HTTPStatus.BAD_REQUEST,
                    f'Reading the uploaded file as an Excel file produced the following'
                    f" <{type(exc).__name__}> error: '{str(exc)}'")

    try:
        config = Config.from_workbook(wbook, sim_hours, num_reps)
    except Exception as exc:
        flask.abort(HTTPStatus.BAD_REQUEST,
                    f'Parsing the uploaded Excel file produced the following'
                    f" <{type(exc).__name__}> error: {str(exc)}")  # openpyxl errors already quoted

    response = new_scenario(config)
    return flask.jsonify(response)


def status(scenario_id: int) -> dict[str, Any]:
    """Return the status of a scenario task."""
    conn = sql.connect(DB_PATH)
    cur = conn.cursor()
    cur.execute(SQL_SELECT_SCENARIO, (scenario_id, ))
    res = cur.fetchone()

    if res is None:
        return None

    _, analysis_id, num_reps, progress, created, completed = res  # unpack tuple
    data = {
        'scenario_id': scenario_id,
        'num_reps': num_reps,
        'progress': progress,
        'created': created,
    }
    if completed is not None:
        data['completed'] = completed
    if analysis_id is not None:
        data['analysis_id'] = analysis_id
    return data


@app.route('/scenarios/<scenario_id>/status/')
@app.route('/scenarios/<scenario_id>/status')
def status_rest(scenario_id) -> Response:
    """Process GET request for querying scenario task status."""
    not_found_text = f"Cannot find scenario with ID: '{scenario_id}'."

    # Ensure scenario_id is integer-compatible
    try:
        s_id = int(scenario_id)
    except ValueError:
        flask.abort(HTTPStatus.NOT_FOUND, description=not_found_text)

    # Fetch the scenario status
    res = status(s_id)
    if res is None:
        flask.abort(HTTPStatus.NOT_FOUND, description=not_found_text)
    return flask.jsonify(res)


@app.route('/scenarios/')
@app.route('/scenarios')
def list_scenarios_rest() -> Response:
    """Return a list of scenarios on the server."""
    conn = sql.connect(DB_PATH)
    df = pd.read_sql(SQL_LIST_SCENARIOS, conn)
    ret = df.to_dict('records')

    # clean up missing values
    for scenario_status in ret:
        if scenario_status['analysis_id'] is None or isnan(scenario_status['analysis_id']):
            del scenario_status['analysis_id']
        if scenario_status['completed'] is None or isnan(scenario_status['completed']):
            del scenario_status['completed']
    return flask.jsonify(ret)


def results_scenario(scenario_id: int) -> str:
    """Return the results of a scenario task."""
    conn = sql.connect(DB_PATH)
    cur = conn.cursor()
    cur.execute(SQL_SCENARIO_RESULTS, (scenario_id, ))
    res = cur.fetchone()
    if res is None or res[0] is None:  # res == None or (None, )
        return None
    return res[0]


@app.route('/scenarios/<scenario_id>/results/')
@app.route('/scenarios/<scenario_id>/results')
def results_scenario_rest(scenario_id: int) -> Response:
    """Process GET request for reading a scenario simulation result."""
    not_found_text = f"Cannot find results for scenario with ID: '{scenario_id}'."

    # Ensure scenario_id is integer-compatible
    try:
        s_id = int(scenario_id)
    except ValueError:
        flask.abort(HTTPStatus.NOT_FOUND, description=not_found_text)

    # Fetch the scenario results
    res = results_scenario(s_id)
    if res is None:
        flask.abort(HTTPStatus.NOT_FOUND, description=not_found_text)

    return app.response_class(res, HTTPStatus.OK, mimetype='application/json')


@app.route('/multi/', methods=['POST'])
@app.route('/multi', methods=['POST'])
def new_multi_rest() -> Response:
    """Process POST request for creating a new multi-scenario analysis.

    The scenario configuration is contained in the request's ``files`` and ``form`` data.
    ``files`` shall be a dict mapping strings to file objects.
    """
    if len(request.files) == 0:
        flask.abort(HTTPStatus.BAD_REQUEST, 'No files submitted for simulation analysis.')

    if 'num_reps' not in request.form:
        flask.abort(HTTPStatus.BAD_REQUEST, 'Missing "num_reps" in request.form.')
    num_reps = int(request.form['num_reps'])  # Number of simulation replications

    if 'sim_hours' not in request.form:
        flask.abort(HTTPStatus.BAD_REQUEST, 'Missing "sim_hours" in request.form.')
    sim_hours = float(request.form['sim_hours'])  # Simulation duration in hours

    # Parse all configs first
    configs: list[Config] = []
    for idx, file in enumerate(request.files.values()):
        try:
            config_bytes = file.stream
            wbook = load_workbook(config_bytes, data_only=True)
            configs.append(Config.from_workbook(wbook, sim_hours, num_reps))
        except Exception as exc:
            flask.abort(
                HTTPStatus.BAD_REQUEST,
                f"Error <{type(exc).__name__}> parsing file #{idx+1} ('{file.name}'): "
                f"{str(exc)}"
            )

    # If all configs valid, create analysis
    conn = sql.connect(DB_PATH)
    cur = conn.cursor()
    cur.execute("""INSERT INTO multis DEFAULT VALUES""")
    conn.commit()
    analysis_id = cur.lastrowid

    # Add the configs to the analysis and enqueue their simulation runs
    for config in configs:
        config.analysis_id = analysis_id
        new_scenario(config)

    return flask.jsonify(status_multi(analysis_id))


def status_multi(analysis_id: int) -> dict[str, Any]:
    """Obtain the status of a multi-scenario analysis."""
    conn = sql.connect(DB_PATH)
    cur = conn.cursor()
    cur.execute(SQL_SELECT_MULTI, (analysis_id, ))
    res = cur.fetchone()
    _, scenario_ids, created, completed, progress = res  # unpack
    scenario_ids = [int(x) for x in scenario_ids.split(',')]
    ret = {
        'analysis_id': analysis_id,
        'scenario_ids': scenario_ids,
        'created': created,
        'progress': progress
    }
    if completed is not None:
        ret['completed'] = completed
    return ret


@app.route('/multi/<analysis_id>/status/')
@app.route('/multi/<analysis_id>/status')
def status_multi_rest(analysis_id) -> Response:
    """Process a GET request for querying multi-scenario analysis status."""
    not_found_text = f"Cannot find analysis with ID: '{analysis_id}'."

    # Ensure scenario_id is integer-compatible
    try:
        a_id = int(analysis_id)
    except ValueError:
        flask.abort(HTTPStatus.NOT_FOUND, description=not_found_text)

    # Fetch the scenario status
    res = status_multi(a_id)
    if res is None:
        flask.abort(HTTPStatus.NOT_FOUND, description=not_found_text)
    return flask.jsonify(res)


@app.route('/multi/')
@app.route('/multi')
def list_multis_rest() -> Response:
    """Return a list of multi-scenario analyses on the server."""
    conn = sql.connect(DB_PATH)
    df = pd.read_sql(SQL_LIST_MULTIS, conn)
    ret = df.to_dict('records')
    for analysis_status in ret:
        analysis_status['scenario_ids'] =\
            [int(x) for x in analysis_status['scenario_ids'].split(',')]
        if analysis_status['completed'] is None or isnan(analysis_status['completed']):
            del analysis_status['completed']
    return flask.jsonify(ret)


@app.route('/multi/<analysis_id>/results/')
@app.route('/multi/<analysis_id>/results')
def results_multi_rest(analysis_id: int) -> Response:
    """Produce a JSON object for a multi-scenario analysis result."""
    not_found_text = f"Cannot find analysis with ID: '{analysis_id}'."
    incomplete_text = f"Analysis with ID: '{analysis_id}' not yet complete."

    # Ensure scenario_id is integer-compatible
    try:
        a_id = int(analysis_id)
    except ValueError:
        flask.abort(HTTPStatus.NOT_FOUND, description=not_found_text)

    # Fetch the scenario status
    res = status_multi(a_id)
    if res is None:
        flask.abort(HTTPStatus.NOT_FOUND, description=not_found_text)
    if res.get('completed') is None:
        flask.abort(HTTPStatus.NOT_FOUND, description=incomplete_text)

    # Fetch each individual result
    all_results = {
        scenario_id: json.loads(results_scenario(scenario_id))
        for scenario_id in res['scenario_ids']
    }
    ret = ({
        'created': res['created'],
        'completed': res['completed'],
        'scenario_ids': res['scenario_ids'],
        'mean_tat': multi_mean_tats(all_results),
        'mean_utilisation': multi_mean_util(all_results),
        'utilisation_hourlies': multi_util_hourlies(all_results)
    })
    return flask.jsonify(ret)


def main() -> None:
    """Start up the Flask server."""

    # Get a SQLite cursor
    conn = sql.connect(DB_PATH)
    cur = conn.cursor()

    cur.execute(SQL_CREATE_MULTIS)
    cur.execute(SQL_CREATE_SCENARIOS)
    conn.commit()

    app.run(host='0.0.0.0', port=BACKEND_PORT, debug=True)


if __name__ == '__main__':
    main()
